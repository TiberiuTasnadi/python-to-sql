from openpyxl import load_workbook
wb = load_workbook(filename='excel.xlsx', read_only=True)
ws = wb['Ultimaker']

def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]
    #for cell in row:
     #   print(cell.value)

hola=list(iter_rows(ws))

dic={"host":'localhost',
            #host=place where is the database
            "user":'root',
            #user = user that has privileges on the db
            "passwd": 'sector7g',
            #password of the user
            "db":'printers'
        }
        db = MySQLdb.connect(**dic)

    #cur is a cursor created to save all the rows
        cur = db.cursor()
    #Now we read each row saved in cur
        
        cur.execute("SELECT %s FROM printers.printers WHERE brand='%s' AND model='%s' ;"%(atr, brand, model))

        for row in cur.fetchall():
            return row[0]
        
        db.commit()
        db.close()
        cur.close()

    except MySQLdb.Error, e:
    #Errors
        print e
